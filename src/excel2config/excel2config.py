'''
Render excel file to text file with Jinja

Version: 2023.02.20
'''

import argparse
import ast
import os
import re
import time
from collections import defaultdict
from itertools import product
from logging import (FileHandler, StreamHandler, basicConfig, error, info,
                     warning)

from jinja2 import Environment, Template, meta
from openpyxl import load_workbook

# LOG OPTIONS
basicConfig(
    handlers=[
        FileHandler('excel2config_LOG.txt'),
        StreamHandler()
    ],
    format='%(asctime)s.%(msecs)03d %(levelname)s : %(message)s',
    datefmt='%Y-%m-%d %H:%M:%S',
    level='INFO'
)


class ExceltoConfig:
    ''' Render excel file to text file with Jinja '''

    def _get_excel_row(self, excel_sheet, row_start=None, row_end=None, col_start=None, col_end=None):
        '''Strip cell and remove none values. 
        excel_sheet is worksheet in openpyxl workbook --> for excel_sheet in load_workbook(filename = 'jinja_excel.xlsx')
        '''
        table_strip = []
        for row in excel_sheet.iter_rows(
                min_row=row_start, max_row=row_end, min_col=col_start, max_col=col_end, values_only=True):
            # 'None' to ''
            row_strip_wo_none = []
            for i in row:
                if str(i) == 'None':
                    row_strip_wo_none.append('')
                else:
                    row_strip_wo_none.append(str(i).strip())

            table_strip.append(row_strip_wo_none)
        return table_strip

    def _create_folder(self, folder_name):
        '''Create folder with timestamp if not exist, return folder name'''
        timestamp = time.strftime("%Y%m%d-%H%M%S")
        output_folder_name = f'{folder_name}_{timestamp}'
        if not os.path.exists(output_folder_name):
            os.makedirs(output_folder_name)
        return output_folder_name

    def _range_text_to_list(self, range_text):
        """ convert range text to list, range must be at the end ("test1/2-3", but not "test1/2-3extra")
        E.g:
            'ip-vrf-1001;mac-vrf-1001'          ==> ['ip-vrf-1001', 'mac-vrf-1001']
            '1/4;2/3-6;3/3/3-5;4;5-6'           ==> ['1/4', '2/3', '2/4', '2/5', '2/6', '3/3/3', '3/3/4', '3/3/5', '4', '5', '6']
            'ip1-2vrf-1001-1005;mac-vrf-6-6'    ==> ['ip1-2vrf-1001', 'ip1-2vrf-1002', 'ip1-2vrf-1003', 'ip1-2vrf-1004', 'ip1-2vrf-1005', 'mac-vrf-6']
        """
        # if no_range [NO_RANGE]
        if range_text.startswith('[NO_RANGE]'):
            return [range_text]
        # if <list in list> python object
        if range_text.startswith('[[') and range_text.endswith(']]'):
            return [range_text]
        # for jinjalist [int1-2;int3]
        if range_text.startswith('[') and range_text.endswith(']'):
            return [range_text]
        # if not range text
        if (';' not in range_text) and (',' not in range_text) and ('-' not in range_text):
            return [range_text]

        range_text_split = range_text.replace(',', ';').split(';')

        result = []
        for i in range_text_split:
            i = i.strip()
            # check empty
            if i == '':
                continue
            # check range exist at last part, e.g. "test1/2-3", but not "test1/3-4extra"
            parts = re.findall('(.*\D|)([0]*)(\d+)-([0]*)(\d+)$', i)
            if parts == []:
                result.append(i)
                continue
            else:
                main_part = parts[0][0]
                range_part_start = parts[0][2]
                range_part_end = parts[0][4]
                # if range start with zero, e.g. port01-03
                range_part_start_zero = parts[0][1]
                range_part_end_zero = parts[0][3]

                for j in range(int(range_part_start), int(range_part_end)+1):
                    # if range start with zero, e.g. port01-03
                    if range_part_start_zero == '' and range_part_end_zero == '':
                        range_result = str(j)
                    else:
                        end_length = len(range_part_end) + \
                            len(range_part_end_zero)
                        range_result = (end_length-len(str(j)))*'0'+str(j)
                    result.append(main_part + range_result)

        return result

    def _get_global_vars(self, input_excel_path):
        wb = load_workbook(filename=input_excel_path, data_only=True)

        # collections defaultdict function
        def defaultdict_():
            return defaultdict(defaultdict_)

        for sheet in wb:
            if sheet.title == 'GLOBAL_VARS':
                all_global = defaultdict_()
                data = self._get_excel_row(sheet, row_start=3, col_start=2)
                for line in data:
                    # if empty line then break
                    if all(v == '' for v in line):
                        break
                    line_range = [self._range_text_to_list(i) for i in line]
                    line_range_product = [i for i in product(*line_range)]
                    for line_product in line_range_product:
                        all_global[line_product[0]
                                   ][line_product[1]] = line_product[2]
                return dict(all_global)
        return {}

    def _get_generate_vars(self, input_excel_path):
        ''' ITER all GLOBAL GEN, call with NEXT '''
        wb = load_workbook(filename=input_excel_path, data_only=True)
        gen_dict = dict()
        for sheet in wb:
            if sheet.title == 'GLOBAL_VARS':
                data = self._get_excel_row(sheet, row_start=3, col_start=2)
                for line in data:
                    if all(v == '' for v in line):
                        break
                    if line[0] == 'GEN':
                        gen_dict[line[1]] = iter(
                            self._range_text_to_list(line[2]))
                return gen_dict
        return {}

    # security problem for "exec" ! restrict import library !
    def _get_func_and_exec(self, input_excel_path):
        ''' Get and Exec Jinja Function '''
        wb = load_workbook(filename=input_excel_path, data_only=True)
        func_dict = dict()
        for sheet in wb:
            if sheet.title == 'GLOBAL_VARS':
                data = self._get_excel_row(sheet, row_start=3, col_start=2)
                for line in data:
                    if all(v == '' for v in line):
                        break
                    if line[0] == 'FUN':
                        line_func_name = line[1]
                        line_func_content = line[2]
                        # # global exec part
                        func_name = line_func_content.split(
                            'def ')[1].split('(')[0]
                        # function name check
                        if func_name != line_func_name:
                            warning(
                                f'FUNCTION NAME <{line_func_name}> DIFFERENT FROM FUNCTION VALUE <{func_name}>')
                        # adding global to def
                        global_text = f"global {func_name}\n"
                        # exec with new function text
                        def_text_exec = global_text + line_func_content
                        def_text_exec_parse = ast.parse(def_text_exec)
                        exec(compile(def_text_exec_parse, "", mode="exec"))
                        # #
                        # add func name to func_dict for jinja render
                        func_dict[line_func_name] = globals()[func_name]
        return func_dict

    def _excel_file_check(self, input_excel_path):
        ''' Input excel file check: size and sheet-variable names (include only "[A-Za-z0-9_]") '''

        try:
            wb = load_workbook(filename=input_excel_path)
            # check if error
            er = False
            for sheet in wb:
                # ignore sheet start with underscore
                if sheet.title[0] == '_':
                    continue

                # CHECK EXCEL SHEET SIZE
                if sheet.max_row > 1000:
                    error(
                        f'MAX 1000 ROW! CHECK EXCEL SHEET REMOVE EMPTY ROW @ {sheet.title}')
                    raise SystemExit
                if sheet.max_column > 1000:
                    error(
                        f'MAX 1000 COLUMN! CHECK EXCEL SHEET REMOVE EMPTY COLUMN @ {sheet.title}')
                    raise SystemExit
                # GLOBAL_VARS sheet check only size
                if sheet.title == 'GLOBAL_VARS':
                    continue

                if not bool(re.match("^[A-Za-z0-9_]*$", sheet.title)) and not sheet.title[0].isdigit():
                    error(
                        f'SHEET NAME! ONLY [A-Za-z0-9_] and NOT START WITH NUMERIC @ {sheet.title}')
                    er = True
                jinja_check = sheet['A2'].value
                for i in re.findall("{{(.+?)}}", jinja_check):
                    # for function
                    # if re.match("^[A-Za-z_]+[A-Za-z0-9_]*\([A-Za-z_]+[A-Za-z0-9_\[\],]*\)", i):
                    if re.match("^[A-Za-z_]+[A-Za-z0-9_]*\(.*\)", i):
                        pass
                    # for list in list python object
                    elif re.match("^[A-Za-z_]+[A-Za-z0-9_]*\[[0-9]+\]", i):
                        pass
                    elif not bool(re.match("^[A-Za-z0-9_]*$", i)) and not i[0].isdigit():
                        error(
                            f'VARIABLE NAME! ONLY [A-Za-z0-9_] and NOT START WITH NUMERIC @ {sheet.title} VARIABLE: {i}')
                        er = True
                header = self._get_excel_row(
                    sheet, row_start=2, col_start=2)[0]
                for i in header:
                    if not bool(re.match("^[A-Za-z0-9_]*$", i)) and not i[0].isdigit():
                        error(
                            f'HEADER NAME! ONLY [A-Za-z0-9_] and NOT START WITH NUMERIC @ {sheet.title} HEADER: {i}')
                        er = True
            if er:
                raise SystemExit
        except Exception as e:
            error(f'EXCEL FILE CHECK NOT DONE! : {e}')
            raise SystemExit

    def run_excel_jinja(self, input_excel_path, output_folder_name_prefix):
        ''' Render jinja host by host with excel sheet data '''
        # check excel file for variable, sheetname and size
        try:
            self._excel_file_check(input_excel_path)
        except Exception as e:
            input('!!! NOT DONE! CHECK ERRORS! Press any key to exit...')
            raise SystemExit
        # create output folder
        try:
            output_folder_name = self._create_folder(output_folder_name_prefix)
            info(f'OUTPUT FOLDER CREATED <{output_folder_name}>')
        except Exception as e:
            error(f'NOT DONE! OUTPUT FOLDER CANNOT CREATED : {e}')
            input('!!! NOT DONE! CHECK ERRORS! Press any key to exit...')
            raise SystemExit
        # run
        wb = load_workbook(filename=input_excel_path, data_only=True)

        # global variable from sheet
        global_vars = self._get_global_vars(input_excel_path)
        gen_vars = self._get_generate_vars(input_excel_path)
        # Jinja function exec and return dict, jinja render for every line
        fun_global_vars = self._get_func_and_exec(input_excel_path)

        for sheet in wb:
            if sheet.title.startswith('_') or sheet.title == 'GLOBAL_VARS':
                continue

            jinja_temp = sheet['A2'].value
            data_check = sheet['B3'].value
            # for removing duplicate logs
            log_duplicate = ''
            # for add header-footer to host once only
            header_footer_host_list = []

            if jinja_temp and data_check:
                info(f'[{input_excel_path}] / [{sheet}] START!')

                # for header-footer create unique host list
                all_host_list_range = [self._range_text_to_list(i[0]) for i in self._get_excel_row(
                    sheet, row_start=3, col_start=2, col_end=2) if i[0] != '']
                all_host_list = set(
                    [j for i in all_host_list_range for j in i])

                # if header append file
                if header_jinja_rgx := re.search(r'{#HEADER\n(.*?)\n#}', jinja_temp, re.DOTALL):
                    header_jinja = header_jinja_rgx.group(1)
                    for i in all_host_list:
                        with open(f'{output_folder_name}/{i}.txt', 'a') as file:
                            file.write(header_jinja+'\n')

                header_w_space = self._get_excel_row(
                    sheet, row_start=2, col_start=2)[0]
                data = self._get_excel_row(sheet, row_start=3, col_start=2)
                header = [i for i in header_w_space if i != '']

                for line in data:
                    # if empty line then break
                    if all(v == '' for v in line):
                        break
                    line_range = [self._range_text_to_list(
                        i) for i in line if i != '']

                    line_range_product = [i for i in product(*line_range)]

                    for line_product in line_range_product:
                        line_render = dict(zip(header, line_product))

                        for key_header in line_render.keys():
                            # if startswith [NO_RANGE], remove [NO_RANGE] and continue without jinja_list
                            if line_render[key_header].startswith('[NO_RANGE]'):
                                line_render[key_header] = line_render[key_header].removeprefix(
                                    '[NO_RANGE]')
                            # if list in list python object
                            elif line_render[key_header].startswith('[[') and line_render[key_header].endswith(']]'):
                                line_render[key_header] = ast.literal_eval(
                                    line_render[key_header])
                            elif line_render[key_header].startswith('[') and line_render[key_header].endswith(']'):
                                # strip and remove bracket first/last
                                removed_bracket = line_render[key_header].strip()[
                                    1:-1]
                                line_render[key_header] = self._range_text_to_list(
                                    removed_bracket)

                        host_name = line_render['host']

                        # check host_global and global vars
                        if 'ALL' in dict(global_vars):
                            all_global_vars = dict(global_vars['ALL'])
                        else:
                            all_global_vars = dict()

                        if host_name in dict(global_vars):
                            host_global_vars = dict(global_vars[host_name])
                        else:
                            host_global_vars = dict()

                        # get jinja template variables
                        jinja_temp_variables = meta.find_undeclared_variables(
                            Environment().parse(jinja_temp))

                        # check jinja_temp has gen_var and add to line_gen_var
                        line_gen_var = {}
                        for gen_var in gen_vars.keys():
                            if gen_var in jinja_temp_variables:
                                try:
                                    line_gen_var[gen_var] = next(
                                        gen_vars[gen_var])
                                except Exception as e:
                                    error(
                                        f'CHECK <{gen_var}> GLOBAL GEN VARIABLE RANGE!')
                                    input(
                                        '!!! NOT DONE! CHECK ERRORS! Press any key to exit...')
                                    raise SystemExit

                        # merge global, host_global and line (priority ordered, last dict replace if same key exist before)
                        # merge line_gen_var
                        host_and_global = (
                            all_global_vars | host_global_vars | line_render | line_gen_var | fun_global_vars)

                        # get jinja template variables and compare with key/value data
                        host_and_global_keys = set(host_and_global.keys())
                        only_jinja_temp_vars = jinja_temp_variables - host_and_global_keys
                        if only_jinja_temp_vars and log_duplicate != sheet.title:
                            warning(
                                f'UNDEFINED VARIABLES FOUND IN JINJA TEMPLATE [{input_excel_path}] / [{sheet}] : <{only_jinja_temp_vars}>')
                            log_duplicate = sheet.title

                        # render with host_and_global and fun_global_vars
                        line_result = Template(
                            jinja_temp).render(host_and_global)
                        line_result += '\n'

                        if line_result.strip():
                            with open(f'{output_folder_name}/{host_name}.txt', 'a') as file:
                                file.write(line_result)

                # if footer append file
                if footer_jinja_rgx := re.search(r'{#FOOTER\n(.*?)\n#}', jinja_temp, re.DOTALL):
                    footer_jinja = footer_jinja_rgx.group(1)
                    for i in all_host_list:
                        with open(f'{output_folder_name}/{i}.txt', 'a') as file:
                            file.write(footer_jinja+'\n\n')

                info(f'[{input_excel_path}] / [{sheet}] DONE!')


def main():
    ''' main function for cli run'''
    parser = argparse.ArgumentParser()
    parser.add_argument(
        'excelfile', help='excel file path [e.g. srlinux_config_1.xlsx] (OPTIONAL, default: config.xlsx)', nargs='?')
    args = parser.parse_args()
    if args.excelfile:
        EXCEL_FILE_PATH = args.excelfile
        excel_file_name = EXCEL_FILE_PATH.split(
            '\\')[-1].split('/')[-1].split('.')[0]
        OUTPUT_FOLDER_NAME_PREFIX = 'OUTPUTS_' + excel_file_name
    else:
        EXCEL_FILE_PATH = 'config.xlsx'
        OUTPUT_FOLDER_NAME_PREFIX = 'OUTPUTS'

    try:
        ExceltoConfig().run_excel_jinja(EXCEL_FILE_PATH, OUTPUT_FOLDER_NAME_PREFIX)
        info('ALL DONE!')
        input('!!! ALL DONE! Press any key to exit...')
    except Exception as e:
        error(f'NOT DONE! CHECK ERRORS : {e}')
        input('!!! NOT DONE! CHECK ERRORS! Press any key to exit...')
        raise SystemExit


if __name__ == "__main__":
    main()
